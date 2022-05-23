from openpyxl import load_workbook, Workbook
from openpyxl.chart import BarChart, Series, Reference
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import datetime
from docxtpl import DocxTemplate

def docgenerator(i):
    # --------------------------
    # ЗАДАНИЕ НАЧАЛЬНЫХ ЗНАЧЕНИЙ
    # --------------------------

    # имя файла
    Name_file_xls = 'BD_act.xlsx'
    Name_file_wd_templ = 'act_template.docx'
    # загрузка файлов
    #wb = load_workbook(Name_file_xls , data_only=True)
    wb = load_workbook(Name_file_xls)
    doc = DocxTemplate(Name_file_wd_templ)
    # загрузка листа
    sch_wb = wb['tb_comp']
    # задам номер строки для взятия данных для генерации докуммента
    number_rows_for_generation_dock = i

    # ------------
    # РАБОЧАЯ ЧАСТЬ
    # -------------

    # соберу данные по ячейкам
    company_full_name = sch_wb.cell(row=number_rows_for_generation_dock, column=1).value
    company_short_name = sch_wb.cell(row=number_rows_for_generation_dock, column=2).value
    contract_number = sch_wb.cell(row=number_rows_for_generation_dock, column=3).value
    contract_date = sch_wb.cell(row=number_rows_for_generation_dock, column=4).value
    number_suppl_agreem = sch_wb.cell(row=number_rows_for_generation_dock, column=5).value
    date_suppl_agreem = sch_wb.cell(row=number_rows_for_generation_dock, column=6).value
    company_seo_short = sch_wb.cell(row=number_rows_for_generation_dock, column=7).value
    job_list = sch_wb.cell(row=number_rows_for_generation_dock, column=8).value
    job_price = sch_wb.cell(row=number_rows_for_generation_dock, column=9).value
    job_price_text = sch_wb.cell(row=number_rows_for_generation_dock, column=10).value
    job_price_nds = sch_wb.cell(row=number_rows_for_generation_dock, column=11).value

    # соберу словарь из подставляемых значений
    context = {
        'company_full_name' : company_full_name,
        'company_short_name' : company_short_name,
        'contract_number' : contract_number,
        'contract_date' : contract_date,
        'number_suppl_agreem' : number_suppl_agreem,
        'date_suppl_agreem' : date_suppl_agreem,
        'company_seo_short' : company_seo_short,
        'job_list' : job_list,
        'job_price' : job_price,
        'job_price_text' : job_price_text,
        'job_price_nds': job_price_nds
    }

    # подставлю значения
    doc.render(context)

    # сохраню в финальный документ
    doc.save("act_final.docx")

docgenerator(2)